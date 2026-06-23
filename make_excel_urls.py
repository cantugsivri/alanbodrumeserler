# -*- coding: utf-8 -*-
import csv, sys, re, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment
sys.stdout.reconfigure(encoding='utf-8')

# --- 1. Yeni eserleri alan_eserler.csv'ye ekle ---
new_artworks = [
    {'id':'90','artwork_name':'VAZO 1','cafe_location':'Sol Duvar','material':'Toprak','dimensions':'18*15','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'91','artwork_name':'VAZO 2','cafe_location':'Sol Duvar','material':'Toprak','dimensions':'18*14','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'92','artwork_name':'VAZO 3','cafe_location':'Sol Duvar','material':'Toprak','dimensions':'19*16','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'93','artwork_name':'PUSULA','cafe_location':'Bar Altı, Yer, Sokak','material':'Taş','dimensions':'55','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'94','artwork_name':'ÇİNGENE KIZ 3','cafe_location':'Sol Duvar','material':'Taş','dimensions':'30*20','price_eur':'50','price_tl':'2500','category':'Mozaik','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'95','artwork_name':'KUŞ 1','cafe_location':'Sol Duvar','material':'Suluboya','dimensions':'25*25','price_eur':'75','price_tl':'4000','category':'Suluboya','artist':'','status':'active','image_url':'','description':'','description_en':''},
    {'id':'96','artwork_name':'KUŞ 2','cafe_location':'Sol Duvar','material':'Suluboya','dimensions':'25*25','price_eur':'75','price_tl':'4000','category':'Suluboya','artist':'','status':'active','image_url':'','description':'','description_en':''},
    {'id':'97','artwork_name':'AT','cafe_location':'Antre','material':'','dimensions':'23,5*40','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'98','artwork_name':'NÜ KADIN 2','cafe_location':'Antre','material':'','dimensions':'30*41','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'99','artwork_name':'YE-İÇ-EĞLEN2','cafe_location':'Sol Duvar','material':'Taş','dimensions':'20*30','price_eur':'','price_tl':'','category':'Mozaik','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
    {'id':'100','artwork_name':'RENKLİ YUVARLAK BALIK','cafe_location':'Antre','material':'','dimensions':'37','price_eur':'','price_tl':'','category':'','artist':'by ÖZGÜR YARDIM','status':'active','image_url':'','description':'','description_en':''},
]

with open('alan_eserler.csv', encoding='utf-8-sig') as f:
    existing = list(csv.DictReader(f))
    fieldnames = list(existing[0].keys())

existing_ids = set(str(r.get('id','')).strip() for r in existing)
to_add = [e for e in new_artworks if e['id'] not in existing_ids]

if to_add:
    with open('alan_eserler.csv', 'a', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        for row in to_add:
            writer.writerow(row)
    print(f'{len(to_add)} yeni eser eklendi.')
else:
    print('Zaten ekli.')

# --- 2. Cloudinary URL'lerini oku ---
foto_by_id = defaultdict(list)
with open('cloudinary_urls.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        eid = row['id'].strip()
        url = row['cloudinary_url'].strip()
        num = row.get('photo_num','').strip()
        fname = row.get('filename','')
        if eid and url:
            if not num or not num.isdigit():
                m = re.search(r'_(\d+)\.\w+$', fname)
                num = m.group(1) if m else '1'
            foto_by_id[eid].append((int(num), url))

for eid in foto_by_id:
    foto_by_id[eid].sort(key=lambda x: x[0])

# --- 3. Tum eserleri oku ---
with open('alan_eserler.csv', encoding='utf-8-sig') as f:
    all_eserler = list(csv.DictReader(f))

all_eserler.sort(key=lambda e: int(str(e.get('id',0)).strip()) if str(e.get('id',0)).strip().isdigit() else 9999)

# --- 4. Excel olustur ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Eser Fotograflari'
ws.append(['eser id no', 'eser ismi', 'url1', 'url2', 'url3'])

header_fill = PatternFill('solid', fgColor='1a1a2e')
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 85
ws.column_dimensions['D'].width = 85
ws.column_dimensions['E'].width = 85

for i, e in enumerate(all_eserler):
    eid = str(e.get('id','')).strip()
    name = e.get('artwork_name','')
    urls = [u for _, u in foto_by_id.get(eid, [])]
    url1 = urls[0] if len(urls) > 0 else ''
    url2 = urls[1] if len(urls) > 1 else ''
    url3 = urls[2] if len(urls) > 2 else ''
    ws.append([eid, name, url1, url2, url3])
    fill = PatternFill('solid', fgColor='f0f4ff' if i % 2 == 0 else 'ffffff')
    for cell in ws[i+2]:
        cell.fill = fill

try:
    wb.save('eser_url_tablosu_v2.xlsx')
    print('Excel hazir: eser_url_tablosu_v2.xlsx')
except PermissionError:
    wb.save('eser_url_tablosu_v3.xlsx')
    print('Excel hazir: eser_url_tablosu_v3.xlsx (v2 dosyasi acik oldugu icin v3 olarak kaydedildi)')
print(f'Toplam: {len(all_eserler)} eser')

eksik = [(e.get('id',''), e.get('artwork_name','')) for e in all_eserler if not foto_by_id.get(str(e.get('id','')))]
print(f'Fotografi olmayan: {len(eksik)}')
for eid, name in eksik:
    print(f'  ID {eid}: {name}')
