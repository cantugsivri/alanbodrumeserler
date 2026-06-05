import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# CSV oku
rows = []
with open('alan_eserler.csv', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    headers = reader.fieldnames
    for row in reader:
        rows.append(row)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ALAN Eser Katalogu"

# ---------- Renkler ----------
DARK       = "1C1B1A"
GOLD       = "C5A059"
CREAM      = "FCFAF6"
LIGHT_GOLD = "F6F0E4"
BORDER_CLR = "E8E3D7"
WHITE      = "FFFFFF"
SECTION_COLORS = {
    "Sol Duvar":            "F0EDE6",
    "Bar Alt\u0131, Yer, Sokak": "EDE8DF",
    "Sa\u011f Duvar":           "E9E4DB",
    "Antre":                "E5E0D6",
    "Tuvalet":              "E2DDD3",
}

# ---------- Başlık satırı (1. satır) ----------
ws.merge_cells("A1:L1")
title_cell = ws["A1"]
title_cell.value = "ALAN ART & COFFEE — Dijital Eser Katalogu"
title_cell.font = Font(name="Garamond", size=16, bold=False, color=WHITE)
title_cell.fill = PatternFill("solid", fgColor=DARK)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ---------- Sütun başlıkları (2. satır) ----------
col_labels = {
    "id":           "No",
    "artwork_name": "Eser Adı",
    "cafe_location":"Konum",
    "material":     "Malzeme",
    "dimensions":   "Ölçüler",
    "price_eur":    "Fiyat (€)",
    "price_tl":     "Fiyat (TL)",
    "category":     "Kategori",
    "artist":       "Sanatçı",
    "status":       "Durum",
    "image_url":    "Fotoğraf URL",
    "description":  "Açıklama",
}
col_widths = [6, 32, 20, 22, 14, 10, 12, 12, 18, 10, 30, 30]

thin_border = Border(
    left=Side(style='thin', color=BORDER_CLR),
    right=Side(style='thin', color=BORDER_CLR),
    top=Side(style='thin', color=BORDER_CLR),
    bottom=Side(style='thin', color=BORDER_CLR),
)

header_fill = PatternFill("solid", fgColor=GOLD)
for col_idx, (key, label) in enumerate(col_labels.items(), start=1):
    cell = ws.cell(row=2, column=col_idx, value=label)
    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

ws.row_dimensions[2].height = 22

# ---------- Veri satırları ----------
current_section = None
for row_idx, row in enumerate(rows, start=3):
    section = row.get("cafe_location", "")
    section_fill_clr = SECTION_COLORS.get(section, CREAM)
    row_fill = PatternFill("solid", fgColor=section_fill_clr if row_idx % 2 == 0 else CREAM)

    for col_idx, key in enumerate(col_labels.keys(), start=1):
        val = row.get(key, "")
        # Sayısal alanlar
        if key in ("id", "price_eur", "price_tl"):
            try:
                val = int(val)
            except (ValueError, TypeError):
                val = 0

        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=(key == "description"))
        cell.font = Font(name="Calibri", size=10, color=DARK)

        # Sayısal hizalama
        if key in ("price_eur", "price_tl"):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = '#,##0'
        if key == "id":
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=10, bold=True, color=GOLD)

    ws.row_dimensions[row_idx].height = 18

# ---------- Dondurulan satırlar ----------
ws.freeze_panes = "A3"

# ---------- Otomatik filtre ----------
ws.auto_filter.ref = f"A2:L{len(rows)+2}"

# ---------- Kaydet ----------
out_path = "ALAN - Eser Katalogu.xlsx"
wb.save(out_path)
print(f"Kaydedildi: {out_path} ({len(rows)} eser)")
